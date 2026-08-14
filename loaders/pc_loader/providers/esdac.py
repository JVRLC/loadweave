"""ESDAC loader — European Soil Database open data (SharePoint Excel files)."""

import argparse
import logging
import os
from io import BytesIO
from typing import Iterator, Optional

import openpyxl
from dotenv import load_dotenv

from loaders.pc_loader.base import BasePCLoader
from loaders.sharepoint_loader.shared.client import SharePointClient

load_dotenv()

logger = logging.getLogger(__name__)

# SharePoint paths relative to site root
FILE_CHEMICAL = (
    "09_DATA/05_Valorisation avancée/01_MYCONNEXION/BDD decrite Mathilde"
    "/Copie de 2020 04 02 open_soil chemical PRA_FR lien avec fichier global.xlsx"
)
FILE_STRUCTURE = (
    "09_DATA/05_Valorisation avancée/01_MYCONNEXION/BDD decrite Mathilde"
    "/Copie de 2020 04 02 open_soil structure PRA_FR lien avec fichier global.xlsx"
)

# Inclusive column counts (0-based, exclusive upper bound)
CHEMICAL_LAST_COL = 13  # A→M
STRUCTURE_LAST_COL = 12  # A→L

IDX_LON = 0
IDX_LAT = 1
FIRST_MEASURE_COL = 2

CATEGORY_MAP = {
    "carbone_c":                          "Chimie",
    "nitrogen_n":                         "Chimie",
    "mo":                                 "Chimie",
    "calcium_carbonates_caco3":           "Chimie",
    "calcium_exchange_capacity_cec":      "Chimie",
    "potassium_k":                        "Chimie",
    "phosphorus_p":                       "Chimie",
    "c_n_ratio":                          "Chimie",
    "ph_in_cacl2_0_01_m_solution":        "Chimie",
    "ph_in_h2o":                          "Chimie",
    "clay_content":                       "Texture",
    "coarse_fragments":                   "Texture",
    "sand_content":                       "Texture",
    "silt_content":                       "Texture",
    "available_water_capacity_awc":       "Texture",
    "bulk_density":                       "Texture",
    "ru_mm":                              "Texture",
    "rfu_mm":                             "Texture",
    "rfu_mo_mm":                          "Texture",
}

# Units keyed by normalized column name; None means dimensionless
UNIT_MAP = {
    "carbone_c": "g.kg-1",
    "nitrogen_n": "g.kg-1",
    "mo": "g.kg-1",
    "calcium_carbonates_caco3": "g.kg-1",
    "calcium_exchange_capacity_cec": "cmol.kg-1",
    "potassium_k": "mg.kg-1",
    "phosphorus_p": "mg.kg-1",
    "clay_content": "%",
    "coarse_fragments": "%",
    "sand_content": "%",
    "silt_content": "%",
    "ru_mm": "mm.cm-1",
    "rfu_mm": "mm.cm-1",
    "rfu_mo_mm": "mm.cm-1",
    "c_n_ratio": None,
    "ph_in_cacl2_0_01_m_solution": None,
    "ph_in_h2o": None,
    "available_water_capacity_awc": None,
    "bulk_density": None,
}


def _normalize(name: str) -> str:
    """Lowercase and replace non-alphanumeric runs with underscores."""
    import re

    return re.sub(r"[^a-z0-9]+", "_", str(name).lower().strip()).strip("_")


class EsdacLoader(BasePCLoader):
    source_name = "ESDAC"
    source_type = "open_source"
    source_description = "European Soil Database — ESDAC (esdac.jrc.ec.europa.eu)"

    def __init__(self):
        self._client = SharePointClient(
            tenant_id=os.getenv("SHAREPOINT_TENANT_ID"),
            client_id=os.getenv("SHAREPOINT_CLIENT_ID"),
            client_secret=os.getenv("SHAREPOINT_CLIENT_SECRET"),
            site_url=os.getenv("SHAREPOINT_SITE_URL"),
        )
        self._client.authenticate()

    def _download(self, file_path: str) -> BytesIO:
        logger.info("Downloading %s...", file_path.split("/")[-1])
        return self._client.download_file(file_path)

    def _parse_headers(self, ws, last_col: int) -> tuple[list[dict], int, int, int]:
        """
        Read header row, find PRA_Code/lon/lat column indices dynamically,
        return (schema, idx_pra, idx_lon, idx_lat).
        """
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        idx_pra = idx_lon = idx_lat = None
        for i, h in enumerate(headers):
            if h is None:
                continue
            norm = _normalize(str(h))
            if norm.endswith("pra") or norm == "pra":
                idx_pra = i
            elif "longitude" in norm:
                idx_lon = i
            elif "latitude" in norm:
                idx_lat = i

        if idx_pra is None or idx_lon is None or idx_lat is None:
            raise RuntimeError(f"PRA_Code/lon/lat columns not found in headers: {headers}")

        meta_cols = {idx_pra, idx_lon, idx_lat}
        schema = []
        for i in range(last_col):
            if i in meta_cols or i >= len(headers) or headers[i] is None:
                continue
            nom = str(headers[i]).strip()
            if not nom:
                continue
            norm = _normalize(nom)
            schema.append({
                "idx":       i,
                "nom":       nom,
                "unite":     UNIT_MAP.get(norm),
                "categorie": CATEGORY_MAP.get(norm),
            })

        logger.info("  pra=col%d, lon=col%d, lat=col%d, %d measure columns",
                    idx_pra, idx_lon, idx_lat, len(schema))
        return schema, idx_pra, idx_lon, idx_lat

    def _parse_file(self, buf: BytesIO, last_col: int) -> dict[str, dict]:
        """Return {pra_code: item} dict from one Excel file."""
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active

        schema, idx_pra, idx_lon, idx_lat = self._parse_headers(ws, last_col)

        items: dict[str, dict] = {}
        skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            pra = row[idx_pra] if idx_pra < len(row) else None
            lon = row[idx_lon] if idx_lon < len(row) else None
            lat = row[idx_lat] if idx_lat < len(row) else None
            if pra is None:
                skipped += 1
                continue

            mesures = []
            for col in schema:
                valeur = row[col["idx"]] if col["idx"] < len(row) else None
                if valeur is None:
                    continue
                mesures.append({
                    "libelle":     col["nom"],
                    "nom_court":   None,
                    "categorie":   col.get("categorie"),
                    "valeur":      valeur,
                    "unite":       col["unite"],
                    "date_mesure": None,
                })

            if not mesures:
                skipped += 1
                continue

            code = str(pra).strip()
            items[code] = {
                "code": code,
                "meta": {
                    "aurea_id": None,
                    "date_prelevement": None,
                    "localisation": f"{lat},{lon}" if lat and lon else None,
                    "culture": None,
                    "type_sol": None,
                    "client_code": None,
                },
                "mesures": mesures,
            }

        if skipped:
            logger.info("  %d rows skipped (missing lon/lat or no measures)", skipped)
        wb.close()
        return items

    def iter_items(self) -> Iterator[dict]:
        """Yield one item per PRA code with measures from both files merged."""
        logger.info("=== ESDAC — chemical file ===")
        chemical = self._parse_file(self._download(FILE_CHEMICAL), CHEMICAL_LAST_COL)

        logger.info("=== ESDAC — structure file ===")
        structure = self._parse_file(self._download(FILE_STRUCTURE), STRUCTURE_LAST_COL)

        all_codes = chemical.keys() | structure.keys()
        for code in all_codes:
            chem = chemical.get(code)
            struct = structure.get(code)
            if chem and struct:
                chem["mesures"] = chem["mesures"] + struct["mesures"]
                yield chem
            else:
                yield chem or struct


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s — %(message)s")

    parser = argparse.ArgumentParser(
        description="ESDAC loader (SharePoint → physico-chemical)"
    )
    parser.parse_args()

    from loaders.pc_loader.runner import run

    logger.info("=== ESDAC loader ===")
    run(EsdacLoader())


if __name__ == "__main__":
    main()
