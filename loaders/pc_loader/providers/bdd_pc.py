import logging
import os
from datetime import date
from io import BytesIO
from typing import Iterator, Optional

import openpyxl
from dotenv import load_dotenv

from loaders.pc_loader.base import BasePCLoader
from loaders.sharepoint_loader.shared.client import SharePointClient

load_dotenv()

logger = logging.getLogger(__name__)

FIRST_MEASURE_COL = 12  # first measure column index (column M)
LAST_MEASURE_COL = (
    100  # exclusive upper bound — stops at column CV (index 99 = SBS par Calcul)
)

# Meta column indices
IDX_TYPE_CONTRAT = 0
IDX_OPP = 1
IDX_CLIENT = 2
IDX_CODE_CLIENT = 3
IDX_PRESTATAIRE = 4
IDX_CODE_ECH = 5
IDX_ANNEE = 6
IDX_PAYS = 7


class BddPCLoader(BasePCLoader):
    source_name = "BDD_PC"
    source_type = "interne"
    source_description = "BDD_PC.xlsx — internal physico-chemical analyses"

    def __init__(self):
        self._client = SharePointClient(
            tenant_id=os.getenv("SHAREPOINT_TENANT_ID"),
            client_id=os.getenv("SHAREPOINT_CLIENT_ID"),
            client_secret=os.getenv("SHAREPOINT_CLIENT_SECRET"),
            site_url=os.getenv("SHAREPOINT_SITE_URL"),
        )
        self._client.authenticate()
        self._file_path = (
            "03_R&D INNOV/30_LABORATOIRE ANALYSES CONTRATS CQ"
            "/02_Physico-chimie - Data & Traçabilité/BDD_PC.xlsx"
        )

    def _download(self) -> BytesIO:
        logger.info("Downloading BDD_PC.xlsx from SharePoint...")
        return self._client.download_file(self._file_path)

    def _parse_schema(self, ws) -> list[dict]:
        """Build measure column schema from the first 3 header rows."""
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        row_cat = rows[0]
        row_unit = rows[1]
        row_head = rows[2]

        schema = []
        for i in range(FIRST_MEASURE_COL, LAST_MEASURE_COL):
            if i >= len(row_head) or row_cat[i] is None:
                continue
            nom = row_head[i]
            if not nom:
                continue
            unite = str(row_unit[i]).strip() if row_unit[i] else None
            if unite == "None":
                unite = None
            # Normalize units with "=" (e.g. "ppm = mg/kg" → "mg/kg")
            if unite and "=" in unite:
                unite = unite.split("=")[-1].strip()
            schema.append(
                {
                    "idx": i,
                    "nom": str(nom).strip(),
                    "categorie": str(row_cat[i]).strip(),
                    "unite": unite,
                }
            )
        return schema

    @staticmethod
    def _to_date(annee) -> Optional[date]:
        try:
            return date(int(annee), 1, 1)
        except (TypeError, ValueError):
            return None

    def _parse_row(self, row: tuple, schema: list[dict]) -> Optional[dict]:
        code = row[IDX_CODE_ECH]
        if not code or str(code).startswith("#"):
            return None

        prestataire = row[IDX_PRESTATAIRE]
        if not prestataire or str(prestataire).startswith("#"):
            return None

        date_prel = self._to_date(row[IDX_ANNEE])

        mesures = []
        for col in schema:
            valeur = row[col["idx"]] if col["idx"] < len(row) else None
            if valeur is None:
                continue
            mesures.append(
                {
                    "libelle":     col["nom"],
                    "nom_court":   None,
                    "categorie":   col.get("categorie"),
                    "valeur":      valeur,
                    "unite":       col["unite"],
                    "date_mesure": date_prel,
                }
            )

        if not mesures:
            return None

        return {
            "code": str(code).strip(),
            "source": {
                "nom": str(prestataire).strip().upper(),
                "type": "prestataire",
                "description": f"Physico-chemical analysis provider ({prestataire})",
            },
            "meta": {
                "aurea_id": None,
                "date_prelevement": date_prel,
                "localisation": str(row[IDX_PAYS]).strip() if row[IDX_PAYS] else None,
                "culture": None,
                "type_sol": (
                    str(row[IDX_TYPE_CONTRAT]).strip()
                    if row[IDX_TYPE_CONTRAT]
                    else None
                ),
                "client_code": (
                    str(row[IDX_CODE_CLIENT]).strip() if row[IDX_CODE_CLIENT] else None
                ),
            },
            "mesures": mesures,
        }

    def iter_items(self) -> Iterator[dict]:
        buf = self._download()
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb["Data"]

        schema = self._parse_schema(ws)
        logger.info("Schema: %d measure columns detected", len(schema))

        seen: set[tuple[str, str]] = set()
        skipped = 0
        duplicates = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            parsed = self._parse_row(row, schema)
            if not parsed:
                skipped += 1
                continue

            key = (parsed["code"], parsed["source"]["nom"])
            if key in seen:
                logger.warning("Duplicate skipped: code=%s prestataire=%s", key[0], key[1])
                duplicates += 1
                continue
            seen.add(key)
            yield parsed

        if skipped:
            logger.info("%d rows skipped (empty code or invalid provider)", skipped)
        if duplicates:
            logger.warning("%d duplicate rows skipped (same code + prestataire)", duplicates)


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s — %(message)s")

    from loaders.pc_loader.runner import run

    logger.info("=== BDD_PC loader ===")
    run(BddPCLoader())


if __name__ == "__main__":
    main()
